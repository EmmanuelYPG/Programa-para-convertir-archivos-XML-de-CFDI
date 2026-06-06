"""
Programa para convertir archivos XML de CFDI (Comprobante Fiscal Digital por Internet) a un archivo Excel (.xlsx) con formato específico para su análisis y reporte.
"""


#! ================================================================
#!                       paqueterias y variables
#! ================================================================
import os
import sys
import threading


def _base_path():
    """Devuelve la carpeta base: temp de extracción si está empaquetado,
    o el directorio del script si se ejecuta directamente."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS          # carpeta temporal de PyInstaller
    try:
        return os.path.dirname(os.path.abspath(__file__))
    except NameError:
        return os.getcwd()
# ──────────────────────────────────────────────────────────────────────────
import xml.etree.ElementTree as ET
from datetime import datetime
from tkinter import (
    Tk, Frame, Label, Button, Entry, Text, Scrollbar, StringVar, BooleanVar,
    Checkbutton, Canvas,
    X, Y, BOTH, LEFT, RIGHT, TOP, BOTTOM, W, END,
    DISABLED, NORMAL, PhotoImage, FLAT
)
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


#* ================================================================
#*               CONSTANTES DE ESTILO (IBC)
#* ================================================================
COLOR_BG            = "#1B1444"   # Fondo principal (morado oscuro)
COLOR_PANEL         = "#262A6B"   # Fondo de paneles
COLOR_PANEL_BORDER  = "#3B3F8F"   # Bordes sutiles
COLOR_ACCENT        = "#E63946"   # Rojo IBC (logo)
COLOR_ACCENT_HOVER  = "#FF5A67"
COLOR_PRIMARY       = "#4A3FB0"   # Azul/morado botón secundario
COLOR_PRIMARY_HOVER = "#5E52C9"
COLOR_TEXT          = "#FFFFFF"
COLOR_MUTED         = "#B8B5D6"
COLOR_INPUT_BG      = "#FFFFFF"
COLOR_INPUT_FG      = "#1B1444"
COLOR_SUCCESS       = "#3FB950"

FONT_FAMILY  = "Dosis"
FONT_TITLE   = (FONT_FAMILY, 22, "bold")
FONT_SUBTITLE= (FONT_FAMILY, 13, "bold")
FONT_LABEL   = (FONT_FAMILY, 11)
FONT_SMALL   = (FONT_FAMILY, 10)
FONT_BUTTON  = (FONT_FAMILY, 12, "bold")
FONT_STATUS  = (FONT_FAMILY, 10, "italic")


#* ================================================================
#*   SAT Nombre espacios de nombres XML para CFDI 3.3 y 4.0, Timbre Fiscal Digital y Nómina 1.2
#* ================================================================
NAMESPACES = {
    'cfdi': 'http://www.sat.gob.mx/cfd/4',
    'cfdi3': 'http://www.sat.gob.mx/cfd/3',
    'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
    'nomina12': 'http://www.sat.gob.mx/nomina12',
    'pago20': 'http://www.sat.gob.mx/Pagos20',
    'pago10': 'http://www.sat.gob.mx/Pagos'
}


#! ================================================================
#!               FUNCIONES DE PARSEO DE CFDI
#! ================================================================
def obtener_valor(elemento, atributo, default=""):
    """
    Obtiene el valor de un atributo de forma segura.
    elemento: el nodo XML del que se quiere obtener el atributo.
    atributo: el nombre del atributo a obtener.
    default: el valor a retornar si el atributo no existe o el elemento es None.
    return: el valor del atributo si existe, o default si no existe o elemento es None.
    """
    return elemento.get(atributo, default) if elemento is not None else default


def parsear_fecha(fecha_str):
    """Extrae solo la fecha de un timestamp ISO."""
    if not fecha_str:
        return ""
    try:
        return fecha_str.split('T')[0]
    except Exception:
        return fecha_str


def extraer_mes(fecha_str):
    """Extrae el mes numérico (MM) de una fecha ISO."""
    if not fecha_str:
        return ""
    try:
        fecha = fecha_str.split('T')[0]
        return fecha.split('-')[1]
    except Exception:
        return ""


def parsear_xml_cfdi(ruta_xml):
    """Extrae todos los datos relevantes de un CFDI XML.
    Devuelve (registro, None) si OK; (None, motivo) si falla."""
    try:
        if os.path.getsize(ruta_xml) == 0:
            return None, "El archivo está vacío (0 bytes)."

        tree = ET.parse(ruta_xml)
        xml_root = tree.getroot()

        version = xml_root.get('Version', '4.0')
        ns = 'cfdi' if version.startswith('4') else 'cfdi3'

        registro = {}

        #? ---- Comprobante ----
        registro['Periodo']             = extraer_mes(xml_root.get('Fecha', ''))
        registro['Version']             = version
        registro['Serie']               = xml_root.get('Serie', '')
        registro['Folio']               = xml_root.get('Folio', '')
        registro['Fecha emision']       = parsear_fecha(xml_root.get('Fecha', ''))
        registro['Tipo']                = xml_root.get('TipoDeComprobante', '')
        registro['SubTotal']            = xml_root.get('SubTotal', '0')
        registro['Descuento']           = xml_root.get('Descuento', '0')
        registro['Total']               = xml_root.get('Total', '0')
        registro['Moneda']              = xml_root.get('Moneda', 'MXN')
        registro['Metodo pago']         = xml_root.get('MetodoPago', '')
        registro['Forma pago']          = xml_root.get('FormaPago', '')
        registro['CP Expedicion']       = xml_root.get('LugarExpedicion', '')
        registro['Exportacion']         = xml_root.get('Exportacion', '')
        registro['Tipo de cambio']      = xml_root.get('TipoCambio', '')
        registro['Condiciones de pago'] = xml_root.get('CondicionesDePago', '')

        #? ---- Emisor ----
        emisor = xml_root.find(f'{{{NAMESPACES[ns]}}}Emisor')
        registro['RFC emisor']     = obtener_valor(emisor, 'Rfc')
        registro['Razon emisor']   = obtener_valor(emisor, 'Nombre')
        registro['Regimen emisor'] = obtener_valor(emisor, 'RegimenFiscal')

        #? ---- Receptor ----
        receptor = xml_root.find(f'{{{NAMESPACES[ns]}}}Receptor')
        registro['RFC receptor']       = obtener_valor(receptor, 'Rfc')
        registro['Razon receptor']     = obtener_valor(receptor, 'Nombre')
        registro['Regimen receptor']   = obtener_valor(receptor, 'RegimenFiscalReceptor')
        registro['Domicilio receptor'] = obtener_valor(receptor, 'DomicilioFiscalReceptor')
        registro['Uso CFDI']           = obtener_valor(receptor, 'UsoCFDI')

        #? ---- Conceptos ----
        conceptos = xml_root.findall(f'.//{{{NAMESPACES[ns]}}}Concepto')
        descripciones = [c.get('Descripcion', '') for c in conceptos if c.get('Descripcion')]
        claves = list({c.get('ClaveProdServ', '') for c in conceptos if c.get('ClaveProdServ')})
        registro['Conceptos']           = ' | '.join(descripciones[:3])
        registro['Claves de productos'] = ', '.join(claves)
        registro['Cuenta Predial']      = ''

        #? ---- Timbre Fiscal ----
        uuid, fecha_cert, pac = '', '', ''
        complemento = xml_root.find(f'{{{NAMESPACES[ns]}}}Complemento')
        if complemento is not None:
            timbre = complemento.find(f'{{{NAMESPACES["tfd"]}}}TimbreFiscalDigital')
            if timbre is not None:
                uuid       = obtener_valor(timbre, 'UUID')
                fecha_cert = parsear_fecha(obtener_valor(timbre, 'FechaTimbrado'))
                pac        = obtener_valor(timbre, 'RfcProvCertif')
        registro['UUID']                = uuid
        registro['Fecha certificacion'] = fecha_cert
        registro['pacCertifico']        = pac

        #? ---- CFDIs relacionados ----
        relaciones = xml_root.find(f'{{{NAMESPACES[ns]}}}CfdiRelacionados')
        if relaciones is not None:
            uuids_rel = [r.get('UUID', '') for r in relaciones.findall(f'{{{NAMESPACES[ns]}}}CfdiRelacionado')]
            registro['Tipo relacion']       = obtener_valor(relaciones, 'TipoRelacion')
            registro['UUIDs relacionados']  = ', '.join(uuids_rel)
        else:
            registro['Tipo relacion']      = ''
            registro['UUIDs relacionados'] = ''

        #? ---- Complemento de Pagos (2.0 / 1.0) ----
        # En CFDI tipo "P" muchos campos viven dentro de pago20:* en lugar de
        # los atributos del Comprobante. Aquí se extraen y se usan como fallback
        # de Forma pago / Tipo de cambio / UUIDs relacionados, además de poblar
        # columnas propias del complemento.
        pago_ns = NAMESPACES['pago20']
        pagos_root = xml_root.find(
            f'.//{{{NAMESPACES[ns]}}}Complemento/{{{pago_ns}}}Pagos'
        )
        if pagos_root is None:
            pago_ns = NAMESPACES['pago10']
            pagos_root = xml_root.find(
                f'.//{{{NAMESPACES[ns]}}}Complemento/{{{pago_ns}}}Pagos'
            )

        fecha_pago = monto_pago = moneda_p = monto_total_pagos = ''
        num_parcialidad = imp_saldo_ant = imp_pagado = imp_saldo_insoluto = ''
        iva16_pago = 0.0
        forma_pago_p = ''
        tipo_cambio_p = ''
        uuids_pago = []

        if pagos_root is not None:
            totales = pagos_root.find(f'{{{pago_ns}}}Totales')
            if totales is not None:
                monto_total_pagos = obtener_valor(totales, 'MontoTotalPagos')
                # Acumulado de IVA 16% trasladado en complemento de pagos
                try:
                    iva16_pago += float(obtener_valor(totales, 'TotalTrasladosImpuestoIVA16', '0') or '0')
                except ValueError:
                    pass

            primer_pago = pagos_root.find(f'{{{pago_ns}}}Pago')
            if primer_pago is not None:
                fecha_pago    = parsear_fecha(obtener_valor(primer_pago, 'FechaPago'))
                monto_pago    = obtener_valor(primer_pago, 'Monto')
                moneda_p      = obtener_valor(primer_pago, 'MonedaP')
                forma_pago_p  = obtener_valor(primer_pago, 'FormaDePagoP')
                tipo_cambio_p = obtener_valor(primer_pago, 'TipoCambioP')

            # Documentos relacionados: tomar del primer pago o, si no, recorrer todos
            for docto in pagos_root.findall(f'.//{{{pago_ns}}}DoctoRelacionado'):
                uuid_dr = obtener_valor(docto, 'IdDocumento')
                if uuid_dr:
                    uuids_pago.append(uuid_dr)
                if not num_parcialidad:
                    num_parcialidad    = obtener_valor(docto, 'NumParcialidad')
                    imp_saldo_ant      = obtener_valor(docto, 'ImpSaldoAnt')
                    imp_pagado         = obtener_valor(docto, 'ImpPagado')
                    imp_saldo_insoluto = obtener_valor(docto, 'ImpSaldoInsoluto')

        # Fallbacks: si el comprobante no trae el atributo, usar el del complemento de pagos
        if not registro['Forma pago'] and forma_pago_p:
            registro['Forma pago'] = forma_pago_p
        if not registro['Tipo de cambio'] and tipo_cambio_p:
            registro['Tipo de cambio'] = tipo_cambio_p
        if not registro['UUIDs relacionados'] and uuids_pago:
            registro['UUIDs relacionados'] = ', '.join(uuids_pago)

        registro['Fecha pago']         = fecha_pago
        registro['Monto pago']         = monto_pago
        registro['Moneda P']           = moneda_p
        registro['Monto total pagos']  = monto_total_pagos
        registro['Num parcialidad']    = num_parcialidad
        registro['Imp saldo anterior'] = imp_saldo_ant
        registro['Imp pagado']         = imp_pagado
        registro['Imp saldo insoluto'] = imp_saldo_insoluto

        #? ---- Impuestos trasladados ----
        iva16 = iva8 = iva0 = iva_exento = 0.0
        ieps8 = ieps0 = 0.0
        impuestos_globales = xml_root.find(f'{{{NAMESPACES[ns]}}}Impuestos')
        if impuestos_globales is not None:
            for traslado in impuestos_globales.findall(f'.//{{{NAMESPACES[ns]}}}Traslado'):
                impuesto = traslado.get('Impuesto', '')
                tasa     = traslado.get('TasaOCuota', '0')
                importe  = float(traslado.get('Importe', '0') or '0')
                if impuesto == '002':
                    if   "0.16" in tasa: iva16 += importe
                    elif "0.08" in tasa: iva8  += importe
                    elif "0.00" in tasa: iva0  += importe
                elif impuesto == '003':
                    if   "0.08" in tasa: ieps8 += importe
                    elif "0.00" in tasa: ieps0 += importe
        # En CFDI tipo P (Pago) los impuestos viven dentro de pago20:Totales,
        # no en cfdi:Impuestos. Sumamos lo que ya recogimos del complemento.
        iva16 += iva16_pago
        registro['IVA Trasladado 16%']  = f"{iva16:.2f}"
        registro['IVA Trasladado 8%']   = f"{iva8:.2f}"
        registro['IVA Trasladado 0%']   = f"{iva0:.2f}"
        registro['IVA Exento']          = f"{iva_exento:.2f}"
        registro['IEPS Trasladado 8%']  = f"{ieps8:.2f}"
        registro['IEPS Trasladado 0%']  = f"{ieps0:.2f}"

        #? ---- Impuestos retenidos ----
        # Solo se leen del bloque global cfdi:Impuestos para no sumar dos veces
        # las retenciones que también aparecen a nivel de cada concepto.
        iva_ret = isr_ret = ieps_ret = 0.0
        if impuestos_globales is not None:
            for retencion in impuestos_globales.findall(f'.//{{{NAMESPACES[ns]}}}Retencion'):
                impuesto = retencion.get('Impuesto', '')
                importe  = float(retencion.get('Importe', '0') or '0')
                if   impuesto == '002': iva_ret  += importe
                elif impuesto == '001': isr_ret  += importe
                elif impuesto == '003': ieps_ret += importe
        registro['IVA Retenido']  = f"{iva_ret:.2f}"
        registro['ISR Retenido']  = f"{isr_ret:.2f}"
        registro['IEPS Retenido'] = f"{ieps_ret:.2f}"

        #? ---- Impuestos locales ----
        registro['Local retenido']           = '0.00'
        registro['Local trasladado (ISH)']   = '0.00'
        registro['Local traslado (ISH)']     = '0.00'

        #? ---- Cancelación ----
        registro['Estado']                        = 'Vigente'
        registro['Fecha proceso cancelacion']     = ''
        registro['Estado cancelacion']            = ''
        registro['Estado proceso cancelacion']    = ''
        registro['Motivo cancelacion']            = ''
        registro['Folio sustitucion cancelacion'] = ''

        #? ---- Campos globales ----
        registro['Global periodicidad'] = ''
        registro['Global meses']        = ''
        registro['Global año']          = ''

        #? ---- Combustibles ----
        registro['SubTotalCombustibles'] = '0.00'
        registro['TotalCombustibles']    = '0.00'

        registro['_ArchivoXml'] = os.path.basename(ruta_xml)

        return registro, None

    except ET.ParseError as e:
        return None, f"XML mal formado: {e}"
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


#! ================================================================
#!                  CONVERSIÓN A EXCEL
#! ================================================================
HEADERS = [
    'Periodo', 'Version', 'UUID', 'UUIDs relacionados', 'Tipo relacion',
    'CP Expedicion', 'Serie', 'Folio', 'Fecha emision', 'Fecha certificacion',
    'pacCertifico', 'Regimen emisor', 'RFC emisor', 'Razon emisor',
    'RFC receptor', 'Razon receptor', 'Regimen receptor', 'Domicilio receptor',
    'Claves de productos', 'Conceptos', 'Cuenta Predial', 'Uso CFDI',
    'Global periodicidad', 'Global meses', 'Global año', 'Condiciones de pago',
    'Estado', 'Fecha proceso cancelacion', 'Estado cancelacion',
    'Estado proceso cancelacion', 'Motivo cancelacion', 'Folio sustitucion cancelacion',
    'Tipo', 'Exportacion', 'Tipo de cambio', 'Metodo pago', 'Forma pago',
    'Moneda', 'Fecha pago', 'Monto pago', 'Moneda P', 'Monto total pagos',
    'Num parcialidad', 'Imp saldo anterior', 'Imp pagado', 'Imp saldo insoluto',
    'SubTotalCombustibles', 'SubTotal', 'Descuento',
    'IVA Trasladado 16%', 'IVA Trasladado 8%', 'IVA Trasladado 0%',
    'IVA Exento', 'IVA Retenido', 'ISR Retenido',
    'IEPS Trasladado 8%', 'IEPS Trasladado 0%', 'IEPS Retenido',
    'Local retenido', 'Local trasladado (ISH)', 'Local traslado (ISH)',
    'TotalCombustibles', 'Total'
]


def convertir_xmls_a_excel(rutas_xml, archivo_salida, progress_cb=None, headers=None):

    """Convierte una lista de rutas XML a un archivo Excel.
    progress_cb(actual, total, mensaje) — opcional, para UI.
    
    rutas_xml: lista de rutas de archivos XML a procesar.
    archivo_salida: ruta completa del archivo Excel a generar.
    progress_cb: función opcional para reportar progreso (actual, total, mensaje).
    headers: lista opcional de columnas a incluir (debe ser subset de HEADERS).
    return: regresa una tupla (exitosos, fallidos, fallidos_info, ruta_excel) con el conteo de resultados y la ruta del Excel generado (o None si no se generó).
    """

    if not headers:
        headers = list(HEADERS)

    wb = Workbook()
    ws = wb.active
    ws.title = "CFDIs"

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
        cell.fill = PatternFill(start_color="262A6B", end_color="262A6B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    exitosos = fallidos = 0
    fallidos_info = []
    total = len(rutas_xml)

    for i, ruta in enumerate(rutas_xml, start=1):
        registro, error = parsear_xml_cfdi(ruta)
        if registro:
            fila = exitosos + 2
            for col, header in enumerate(headers, start=1):
                ws.cell(row=fila, column=col).value = registro.get(header, '')
            exitosos += 1
        else:
            fallidos += 1
            fallidos_info.append((os.path.basename(ruta), error or "motivo desconocido"))

        if progress_cb:
            progress_cb(i, total, os.path.basename(ruta))

    if exitosos == 0:
        return exitosos, fallidos, fallidos_info, None

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_length:
                    max_length = len(v)
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(archivo_salida)
    return exitosos, fallidos, fallidos_info, archivo_salida


def buscar_xmls_en_carpeta(carpeta):
    """Recorre recursivamente una carpeta y devuelve las rutas de los XML encontrados."""
    encontrados = []
    for dp, _, files in os.walk(carpeta):
        for f in files:
            if f.lower().endswith('.xml'):
                encontrados.append(os.path.join(dp, f))
    return encontrados


#! ================================================================
#!                    Interfaz visual
#! ================================================================
class ConverterApp:
    def __init__(self, root):

        """
        Esta función inicializa la aplicación, configurando la ventana principal y preparando las variables necesarias.
        root: instancia de Tk() para la ventana principal.
        """

        self.root = root
        self.ruta_entrada = StringVar()
        self.ruta_salida  = StringVar()
        self.archivos_seleccionados = []
        self.modo_seleccion = "carpeta"
        self._converting = False

        self._configurar_ventana()
        self._cargar_logo()
        self._construir_ui()

    #? ---------- Configuración de ventana ----------
    def _configurar_ventana(self):

        """
        Esta función:
        Configura las propiedades de la ventana principal, como el título, tamaño, color de fondo y posición centrada en la pantalla.
        """

        self.root.title("Convertidor de XMLs a Excel — IBC")
        w, h = 1280, 720            # Tamaño de la intefaz
        self.root.geometry(f"{w}x{h}")
        self.root.resizable(False, False)
        self.root.configure(bg=COLOR_BG)

        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _cargar_logo(self):

        """
        Busca un archivo de logo en el directorio del script y lo usa como
        favicon y dentro del header. Si no existe, la app sigue funcionando.
        """

        self.logo_img = None
        self.logo_header = None
        base = _base_path()

        for nombre in ("IBC-Logo-2026_sin_fondo.png", "logo.png", "ibc_logo.png", "logo.gif", "ibc.png"):
            ruta = os.path.join(base, nombre)
            if os.path.exists(ruta):
                try:
                    self.logo_img = PhotoImage(file=ruta)
                    self.root.iconphoto(True, self.logo_img)
                    target_h = 60
                    factor = max(1, self.logo_img.height() // target_h)
                    self.logo_header = self.logo_img.subsample(factor, factor)
                except Exception as e:
                    print(f"No se pudo cargar el logo: {e}")
                break

    #? ---------- Construcción de UI ----------
    def _construir_ui(self):

        """
        Esta función construye toda la interfaz de usuario, organizando los diferentes paneles y componentes.
        Colorcitos, botones, entradas, panel lateral, status, etc. Todo lo que el usuario ve y con lo que interactúa se crea aquí.
        """

        self._construir_header()

        # Footer fijo abajo y cuerpo (con split izquierdo/derecho) llenando el resto
        self._construir_footer()

        body = Frame(self.root, bg=COLOR_BG)
        body.pack(side=TOP, fill=BOTH, expand=True)

        self.main_left = Frame(body, bg=COLOR_BG)
        self.main_left.pack(side=LEFT, fill=BOTH, expand=True)

        self.main_right = Frame(body, bg=COLOR_BG, width=300)
        self.main_right.pack(side=RIGHT, fill=Y, padx=(0, 25), pady=(18, 10))
        self.main_right.pack_propagate(False)

        self._construir_panel_entrada()
        self._construir_panel_salida()
        self._construir_accion()
        self._construir_status()
        self._construir_panel_columnas()

    def _construir_header(self):
        header = Frame(self.root, bg=COLOR_BG)
        header.pack(fill=X, padx=25, pady=(18, 8))

        if self.logo_header is not None:
            Label(header, image=self.logo_header, bg=COLOR_BG).pack(side=LEFT, padx=(0, 18))

        titulo_frame = Frame(header, bg=COLOR_BG)
        titulo_frame.pack(side=LEFT, anchor=W)

        Label(
            titulo_frame,
            text="Convertidor de XMLs a Excel",
            font=FONT_TITLE,
            bg=COLOR_BG,
            fg=COLOR_TEXT
        ).pack(anchor=W)

        Label(
            titulo_frame,
            text="Innovation, Business & Consulting",
            font=FONT_SMALL,
            bg=COLOR_BG,
            fg=COLOR_MUTED
        ).pack(anchor=W)

        Frame(self.root, bg=COLOR_ACCENT, height=2).pack(fill=X, padx=25)

    def _construir_panel_entrada(self):
        panel = Frame(self.main_left, bg=COLOR_PANEL,
                        highlightbackground=COLOR_PANEL_BORDER, highlightthickness=1)
        panel.pack(fill=X, padx=25, pady=(18, 10))

        Label(
            panel,
            text="1.  Seleccionar archivos XML",
            font=FONT_SUBTITLE,
            bg=COLOR_PANEL,
            fg=COLOR_TEXT
        ).pack(anchor=W, padx=18, pady=(14, 4))

        Label(
            panel,
            text="Elige la carpeta que contiene los XMLs (busca en subcarpetas) o "
                    "selecciona los archivos individualmente.",
            font=FONT_SMALL,
            bg=COLOR_PANEL,
            fg=COLOR_MUTED,
            wraplength=880,
            justify=LEFT
        ).pack(anchor=W, padx=18, pady=(0, 10))

        fila = Frame(panel, bg=COLOR_PANEL)
        fila.pack(fill=X, padx=18, pady=(0, 10))

        Entry(
            fila,
            textvariable=self.ruta_entrada,
            font=FONT_LABEL,
            bg=COLOR_INPUT_BG,
            fg=COLOR_INPUT_FG,
            relief=FLAT,
            state="readonly",
            readonlybackground=COLOR_INPUT_BG
        ).pack(side=LEFT, fill=X, expand=True, ipady=6, padx=(0, 10))

        self._boton(fila, "Examinar carpeta", self._seleccionar_carpeta_entrada,
                    COLOR_PRIMARY, COLOR_PRIMARY_HOVER).pack(side=LEFT, padx=(0, 6))

        self._boton(fila, "Elegir archivos", self._seleccionar_archivos,
                    COLOR_PRIMARY, COLOR_PRIMARY_HOVER).pack(side=LEFT)

        self.lbl_conteo = Label(
            panel,
            text="Ningún archivo seleccionado aún.",
            font=FONT_SMALL,
            bg=COLOR_PANEL,
            fg=COLOR_MUTED
        )
        self.lbl_conteo.pack(anchor=W, padx=18, pady=(0, 14))

    def _construir_panel_salida(self):
        panel = Frame(self.main_left, bg=COLOR_PANEL,
                    highlightbackground=COLOR_PANEL_BORDER, highlightthickness=1)
        panel.pack(fill=X, padx=25, pady=(0, 10))

        Label(
            panel,
            text="2.  Carpeta de guardado",
            font=FONT_SUBTITLE,
            bg=COLOR_PANEL,
            fg=COLOR_TEXT
        ).pack(anchor=W, padx=18, pady=(14, 4))

        Label(
            panel,
            text="Indica la carpeta donde se generará el archivo Excel (.xlsx) con los datos consolidados.",
            font=FONT_SMALL,
            bg=COLOR_PANEL,
            fg=COLOR_MUTED,
            wraplength=880,
            justify=LEFT
        ).pack(anchor=W, padx=18, pady=(0, 10))

        fila = Frame(panel, bg=COLOR_PANEL)
        fila.pack(fill=X, padx=18, pady=(0, 14))

        Entry(
            fila,
            textvariable=self.ruta_salida,
            font=FONT_LABEL,
            bg=COLOR_INPUT_BG,
            fg=COLOR_INPUT_FG,
            relief=FLAT,
            state="readonly",
            readonlybackground=COLOR_INPUT_BG
        ).pack(side=LEFT, fill=X, expand=True, ipady=6, padx=(0, 10))

        self._boton(fila, "Examinar", self._seleccionar_carpeta_salida,
                    COLOR_PRIMARY, COLOR_PRIMARY_HOVER).pack(side=LEFT)

    def _construir_accion(self):
        contenedor = Frame(self.main_left, bg=COLOR_BG)
        contenedor.pack(fill=X, padx=25, pady=(8, 6))

        self.btn_convertir = self._boton(
            contenedor,
            "Convertir a Excel",
            self._iniciar_conversion,
            COLOR_ACCENT,
            COLOR_ACCENT_HOVER,
            ancho=22,
            alto=2,
            font=FONT_BUTTON
        )
        self.btn_convertir.pack(pady=4)

        estilo = ttk.Style()
        try:
            estilo.theme_use('clam')
        except Exception:
            pass
        estilo.configure(
            "IBC.Horizontal.TProgressbar",
            troughcolor=COLOR_PANEL,
            background=COLOR_ACCENT,
            bordercolor=COLOR_PANEL,
            lightcolor=COLOR_ACCENT,
            darkcolor=COLOR_ACCENT,
            thickness=14
        )
        self.progreso = ttk.Progressbar(
            self.main_left,
            style="IBC.Horizontal.TProgressbar",
            mode="determinate",
            length=800
        )
        self.progreso.pack(padx=25, pady=(10, 4))

    def _construir_status(self):
        self.lbl_status = Label(
            self.main_left,
            text="Listo.",
            font=FONT_STATUS,
            bg=COLOR_BG,
            fg=COLOR_MUTED,
            anchor=W
        )
        self.lbl_status.pack(fill=X, padx=28, pady=(4, 4))

        #? --- Panel de archivos fallidos (siempre visible, se llena tras convertir) ---
        cont_fallidos = Frame(self.main_left, bg=COLOR_BG)
        cont_fallidos.pack(fill=X, padx=25, pady=(0, 8))

        self.lbl_fallidos_titulo = Label(
            cont_fallidos,
            text="Archivos fallidos:",
            font=FONT_SMALL,
            bg=COLOR_BG,
            fg=COLOR_MUTED,
            anchor=W
        )
        self.lbl_fallidos_titulo.pack(fill=X, padx=3, pady=(0, 2))

        caja = Frame(cont_fallidos, bg=COLOR_PANEL_BORDER,
                    highlightbackground=COLOR_PANEL_BORDER, highlightthickness=1)
        caja.pack(fill=X)

        scroll = Scrollbar(caja)
        scroll.pack(side=RIGHT, fill=Y)

        self.txt_fallidos = Text(
            caja,
            height=4,
            bg=COLOR_PANEL,
            fg=COLOR_TEXT,
            font=FONT_SMALL,
            relief=FLAT,
            bd=0,
            yscrollcommand=scroll.set,
            wrap="word",
            padx=8,
            pady=6
        )
        self.txt_fallidos.pack(side=LEFT, fill=X, expand=True)
        scroll.config(command=self.txt_fallidos.yview)

        self.txt_fallidos.tag_configure("err", foreground=COLOR_ACCENT)
        self.txt_fallidos.tag_configure("hint", foreground=COLOR_MUTED)
        self._mostrar_fallidos_vacio()

    def _construir_footer(self):
        Label(
            self.root,
            text="© IBC — Innovation, Business & Consulting",
            font=FONT_SMALL,
            bg=COLOR_BG,
            fg=COLOR_MUTED
        ).pack(side=BOTTOM, pady=(0, 10))

    def _construir_panel_columnas(self):
        """Panel lateral derecho con un Checkbutton por cada columna del Excel.
        El usuario puede deseleccionar las que no desea exportar."""
        panel = Frame(self.main_right, bg=COLOR_PANEL,
                    highlightbackground=COLOR_PANEL_BORDER, highlightthickness=1)
        panel.pack(fill=BOTH, expand=True)

        Label(
            panel,
            text="3.  Columnas a exportar",
            font=FONT_SUBTITLE,
            bg=COLOR_PANEL,
            fg=COLOR_TEXT
        ).pack(anchor=W, padx=14, pady=(14, 4))

        Label(
            panel,
            text="Marca o desmarca las columnas que se incluirán en el Excel.",
            font=FONT_SMALL,
            bg=COLOR_PANEL,
            fg=COLOR_MUTED,
            wraplength=260,
            justify=LEFT
        ).pack(anchor=W, padx=14, pady=(0, 8))

        # Botones "Todos" y "Ninguno" para seleccionar/deseleccionar todas las columnas rápidamente
        fila_btns = Frame(panel, bg=COLOR_PANEL)
        fila_btns.pack(fill=X, padx=14, pady=(0, 8))
        self._boton(fila_btns, "Todos", self._seleccionar_todas_columnas,
                    COLOR_PRIMARY, COLOR_PRIMARY_HOVER, ancho=10).pack(side=LEFT, padx=(0, 6))
        self._boton(fila_btns, "Ninguno", self._deseleccionar_todas_columnas,
                    COLOR_PRIMARY, COLOR_PRIMARY_HOVER, ancho=10).pack(side=LEFT)

        # Contenedor scrollable
        cont = Frame(panel, bg=COLOR_PANEL)
        cont.pack(fill=BOTH, expand=True, padx=10, pady=(0, 12))

        canvas = Canvas(cont, bg=COLOR_PANEL, highlightthickness=0, bd=0)
        scroll = Scrollbar(cont, orient="vertical", command=canvas.yview)
        inner = Frame(canvas, bg=COLOR_PANEL)

        inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)

        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scroll.pack(side=RIGHT, fill=Y)

        # Scroll con la rueda del mouse solo cuando el cursor está sobre el panel
        def _on_wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Un Checkbutton por cada header, todos en True por defecto
        self.columnas_vars = {}
        for header in HEADERS:
            var = BooleanVar(value=True)
            cb = Checkbutton(
                inner,
                text=header,
                variable=var,
                onvalue=True,
                offvalue=False,
                bg=COLOR_PANEL,
                fg=COLOR_TEXT,
                activebackground=COLOR_PANEL,
                activeforeground=COLOR_TEXT,
                selectcolor=COLOR_BG,
                font=FONT_SMALL,
                anchor=W,
                highlightthickness=0,
                bd=0,
                padx=2,
                pady=1,
                cursor="hand2"
            )
            cb.pack(fill=X, anchor=W)
            self.columnas_vars[header] = var

    def _seleccionar_todas_columnas(self):
        for var in self.columnas_vars.values():
            var.set(True)

    def _deseleccionar_todas_columnas(self):
        for var in self.columnas_vars.values():
            var.set(False)

    def _mostrar_fallidos_vacio(self):
        self.txt_fallidos.config(state=NORMAL)
        self.txt_fallidos.delete("1.0", END)
        self.txt_fallidos.insert(END, "Aún no se ha ejecutado ninguna conversión.", "hint")
        self.txt_fallidos.config(state=DISABLED)

    def _mostrar_fallidos(self, fallidos_info):
        self.txt_fallidos.config(state=NORMAL)
        self.txt_fallidos.delete("1.0", END)
        if not fallidos_info:
            self.txt_fallidos.insert(END, "Sin archivos fallidos en esta conversión.", "hint")
        else:
            for nombre, motivo in fallidos_info:
                self.txt_fallidos.insert(END, f"• {nombre}", "err")
                self.txt_fallidos.insert(END, f"  —  {motivo}\n", "hint")
        self.txt_fallidos.config(state=DISABLED)

    def _boton(self, parent, texto, comando, color, color_hover,
                ancho=16, alto=1, font=None):
        """
        Crea un botón con estilo IBC y efecto hover.
        """
        btn = Button(
            parent,
            text=texto,
            command=comando,
            bg=color,
            fg=COLOR_TEXT,
            activebackground=color_hover,
            activeforeground=COLOR_TEXT,
            font=font or FONT_LABEL,
            relief=FLAT,
            cursor="hand2",
            width=ancho,
            height=alto,
            bd=0
        )
        btn.bind("<Enter>", lambda e: btn.config(bg=color_hover))
        btn.bind("<Leave>", lambda e: btn.config(bg=color))
        return btn

    def _seleccionar_carpeta_entrada(self):
        """
        Abre un diálogo para seleccionar una carpeta, busca XMLs dentro y actualiza el estado.
        :param self: instancia de la aplicación
        :return: None
        """

        carpeta = filedialog.askdirectory(title="Selecciona la carpeta con los XMLs")
        if not carpeta:
            return
        self.modo_seleccion = "carpeta"
        self.ruta_entrada.set(carpeta)
        self.archivos_seleccionados = buscar_xmls_en_carpeta(carpeta)
        n = len(self.archivos_seleccionados)
        if n == 0:
            self.lbl_conteo.config(
                text="No se encontraron archivos XML en la carpeta seleccionada.",
                fg=COLOR_ACCENT
            )
        else:
            self.lbl_conteo.config(
                text=f"{n} archivo(s) XML encontrado(s) en la carpeta.",
                fg=COLOR_SUCCESS
            )

    def _seleccionar_archivos(self):
        """
        Abre un diálogo para seleccionar archivos XML individualmente y actualiza el estado.
        :param self: instancia de la aplicación
        :return: None
        """

        archivos = filedialog.askopenfilenames(
            title="Selecciona archivos XML",
            filetypes=[("Archivos XML", "*.xml"), ("Todos los archivos", "*.*")]
        )
        if not archivos:
            return
        self.modo_seleccion = "archivos"
        self.archivos_seleccionados = list(archivos)
        self.ruta_entrada.set(f"{len(archivos)} archivo(s) seleccionado(s)")
        self.lbl_conteo.config(
            text=f"{len(archivos)} archivo(s) XML seleccionado(s) manualmente.",
            fg=COLOR_SUCCESS
        )

    def _seleccionar_carpeta_salida(self):
        carpeta = filedialog.askdirectory(title="Selecciona la carpeta de guardado")
        if carpeta:
            self.ruta_salida.set(carpeta)

    def _iniciar_conversion(self):
        """
        Valida las rutas seleccionadas, muestra advertencias si falta algo, y si todo está listo, inicia la conversión en un hilo separado.
        :param self: instancia de la aplicación
        :return: None
        """

        if self._converting:
            return

        if not self.archivos_seleccionados:
            messagebox.showwarning(
                "Falta selección",
                "Primero debes seleccionar una carpeta o archivos XML."
            )
            return

        if not self.ruta_salida.get():
            messagebox.showwarning(
                "Falta carpeta de salida",
                "Selecciona la carpeta donde se guardará el archivo Excel."
            )
            return

        if not os.path.isdir(self.ruta_salida.get()):
            messagebox.showerror("Carpeta inválida", "La carpeta de salida no existe.")
            return

        headers_seleccionados = [h for h in HEADERS if self.columnas_vars[h].get()]
        if not headers_seleccionados:
            messagebox.showwarning(
                "Sin columnas seleccionadas",
                "Debes seleccionar al menos una columna a exportar."
            )
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_salida = os.path.join(
            self.ruta_salida.get(),
            f"Reporte_CFDIs_{timestamp}.xlsx"
        )

        self._converting = True
        self.btn_convertir.config(state=DISABLED, text="Procesando...")
        self.progreso['value'] = 0
        self.progreso['maximum'] = len(self.archivos_seleccionados)
        self.lbl_status.config(text="Iniciando conversión...", fg=COLOR_MUTED)
        self._mostrar_fallidos([])
        self.lbl_fallidos_titulo.config(
            text="Archivos fallidos:", fg=COLOR_MUTED
        )

        hilo = threading.Thread(
            target=self._ejecutar_conversion,
            args=(list(self.archivos_seleccionados), archivo_salida, headers_seleccionados),
            daemon=True
        )
        hilo.start()

    def _ejecutar_conversion(self, rutas, archivo_salida, headers):
        def progress_cb(actual, total, nombre):
            self.root.after(0, self._actualizar_progreso, actual, total, nombre)

        try:
            exitosos, fallidos, fallidos_info, archivo_generado = \
                convertir_xmls_a_excel(rutas, archivo_salida, progress_cb, headers)
            self.root.after(0, self._conversion_terminada,
                            True, exitosos, fallidos, fallidos_info,
                            archivo_generado, None)
        except Exception as e:
            self.root.after(0, self._conversion_terminada,
                            False, 0, 0, [], None, str(e))

    def _actualizar_progreso(self, actual, total, nombre):
        self.progreso['value'] = actual
        self.lbl_status.config(
            text=f"Procesando {actual}/{total} — {nombre}",
            fg=COLOR_MUTED
        )

    def _conversion_terminada(self, ok, exitosos, fallidos, fallidos_info, archivo, error):
        self._converting = False
        self.btn_convertir.config(state=NORMAL, text="Convertir a Excel")

        if not ok:
            self.lbl_status.config(text=f"Error: {error}", fg=COLOR_ACCENT)
            messagebox.showerror("Error en la conversión", f"Ocurrió un error:\n{error}")
            return

        self.progreso['value'] = self.progreso['maximum']

        # Mostrar siempre el listado de fallidos
        self._mostrar_fallidos(fallidos_info)
        if fallidos:
            self.lbl_fallidos_titulo.config(
                text=f"Archivos fallidos ({fallidos}):", fg=COLOR_ACCENT
            )
        else:
            self.lbl_fallidos_titulo.config(
                text="Archivos fallidos:", fg=COLOR_MUTED
            )

        # Caso: ningún XML válido 
        if archivo is None:
            self.lbl_status.config(
                text=f"No se generó el Excel: 0 registros exitosos, {fallidos} fallidos.",
                fg=COLOR_ACCENT
            )
            messagebox.showwarning(
                "Sin registros válidos",
                f"No se pudo procesar ningún XML ({fallidos} fallidos).\n\n"
                "No se generó el archivo Excel.\n"
                "Revisa el detalle de los archivos fallidos en la parte inferior."
            )
            return

        resumen = f"Exitosos: {exitosos}  |  Fallidos: {fallidos}"
        self.lbl_status.config(text=f"Conversión completada.  {resumen}", fg=COLOR_SUCCESS)

        msg = (
            f"Se procesaron {exitosos} registros exitosamente"
            + (f" y {fallidos} fallidos." if fallidos else ".")
            + f"\n\nArchivo generado:\n{archivo}\n\n"
        )
        if fallidos:
            msg += "Revisa los archivos fallidos en la parte inferior.\n\n"
        msg += "¿Deseas abrir la carpeta de salida?"

        if messagebox.askyesno("Conversión completada", msg):
            try:
                if sys.platform.startswith("win"):
                    os.startfile(os.path.dirname(archivo))
                elif sys.platform == "darwin":
                    os.system(f'open "{os.path.dirname(archivo)}"')
                else:
                    os.system(f'xdg-open "{os.path.dirname(archivo)}"')
            except Exception:
                pass


#! ================================================================
#!   FUNCIÓN PRINCIPAL de la aplicación para que no se cierre
#! ================================================================
def main():
    root = Tk()
    ConverterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
